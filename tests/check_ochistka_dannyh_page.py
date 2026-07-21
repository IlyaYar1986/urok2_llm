import unittest
from html.parser import HTMLParser
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


ROOT = Path(__file__).resolve().parents[1]
PAGE = ROOT / "web" / "ochistka_dannyh.html"
PREV_PAGE = ROOT / "web" / "cifrovoy_sled.html"
SAMPLE_XLSX = ROOT / "files" / "uspevaemost_gruppy_primer.xlsx"
VIZ_PAGES = [
    ROOT / "web" / "gryaznye_dannye.html",
    ROOT / "web" / "trenazher_ochistki.html",
    ROOT / "web" / "ii_vizualizator.html",
]


class PageParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.ids = set()
        self.id_counts = {}
        self.iframe_srcs = []
        self.img_srcs = []

    def handle_starttag(self, tag, attrs):
        values = dict(attrs)
        if "id" in values:
            value = values["id"]
            self.ids.add(value)
            self.id_counts[value] = self.id_counts.get(value, 0) + 1
        if tag == "iframe" and "src" in values:
            self.iframe_srcs.append(values["src"])
        if tag == "img" and "src" in values:
            self.img_srcs.append(values["src"])


class OchistkaDannyhPageTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.html = PAGE.read_text(encoding="utf-8") if PAGE.exists() else ""
        cls.parser = PageParser()
        cls.parser.feed(cls.html)

    def test_page_exists_with_required_sections(self):
        self.assertTrue(PAGE.is_file())
        required = {
            "hero",
            "s1",
            "s2",
            "s3",
            "s4",
            "s5",
            "checklist",
            "quiz",
            "final",
        }
        self.assertTrue(required.issubset(self.parser.ids))
        duplicates = [
            element_id
            for element_id, count in self.parser.id_counts.items()
            if count > 1
        ]
        self.assertEqual([], duplicates)

    def test_lecture_markers_are_present(self):
        for marker in (
            "рязные данные",
            "Анонимизация",
            "тандартизация",
            "труктурирование",
            "Ctrl+H",
            "Present Simple",
            "Claude",
            "152-ФЗ",
            "архитектор данных",
        ):
            self.assertIn(marker, self.html)

    def test_services_have_active_links(self):
        for url in (
            "https://presentsimple.ai/",
            "https://claude.ai/",
            "https://giga.chat/",
            "https://chat.deepseek.com/",
        ):
            self.assertIn(url, self.html)

    def test_sample_xlsx_is_downloadable_and_valid(self):
        self.assertIn("../files/uspevaemost_gruppy_primer.xlsx", self.html)
        self.assertTrue(SAMPLE_XLSX.is_file())
        if openpyxl is None:
            self.skipTest("openpyxl not installed")
        workbook = openpyxl.load_workbook(SAMPLE_XLSX)
        sheet = workbook.active
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertIn("ID Студента", header)
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertGreaterEqual(len(rows), 5)
        for row in rows:
            self.assertTrue(str(row[0]).startswith("Студент"))
            self.assertNotIn(None, row)

    def test_visualizations_are_embedded_and_exist(self):
        expected = {
            "gryaznye_dannye.html",
            "trenazher_ochistki.html",
            "ii_vizualizator.html",
        }
        embedded = {
            src for src in self.parser.iframe_srcs if not src.startswith("#")
        }
        self.assertTrue(expected.issubset(embedded))
        for viz_page in VIZ_PAGES:
            self.assertTrue(viz_page.is_file(), viz_page.name)

    def test_infographics_point_to_ochistka_dannyh_folder(self):
        png_srcs = [
            src for src in self.parser.img_srcs if src.endswith(".png")
        ]
        self.assertGreaterEqual(len(png_srcs), 6)
        for src in png_srcs:
            self.assertTrue(
                src.startswith("../png/ochistka_dannyh/"),
                src,
            )

    def test_viz_pages_are_self_contained(self):
        for viz_page in VIZ_PAGES:
            html = viz_page.read_text(encoding="utf-8")
            self.assertIn('lang="ru"', html, viz_page.name)
            for forbidden in ("http://", "https://", "cdn.", "unpkg", "jsdelivr"):
                self.assertNotIn(forbidden, html, viz_page.name)

    def test_previous_lecture_links_forward(self):
        prev_html = PREV_PAGE.read_text(encoding="utf-8")
        self.assertIn("ochistka_dannyh.html", prev_html)
        self.assertIn("cifrovoy_sled.html", self.html)


if __name__ == "__main__":
    unittest.main()
